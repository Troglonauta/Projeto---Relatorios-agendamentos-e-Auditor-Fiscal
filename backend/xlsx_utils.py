"""Helpers de geracao de XLSX formatado — Sprint 15.

Padroniza a aparencia oficial dos relatorios da Fertimaxi:
- Header verde Fertimaxi (#2E8B3D) com texto branco bold centralizado.
- AutoFit por amostragem (cap em 60 caracteres).
- Linha 1 congelada (freeze panes A2).
- AutoFilter na linha de header.
- Reordenacao de colunas pelo prefixo da tabela (SF1 -> SD1 -> SDS -> SDT).

Uso:
    from backend.xlsx_utils import build_formatted_xlsx_bytes
    blob = build_formatted_xlsx_bytes(rows, sheet_name="Anomalias")
    # blob e' bytes — pode ir para FileResponse, anexo de email, etc.
"""
from __future__ import annotations

import io
from typing import Iterable, Optional, Sequence

# Cores oficiais Fertimaxi
HEADER_BG_HEX  = "FF2E8B3D"   # verde principal
HEADER_FG_HEX  = "FFFFFFFF"   # branco
WIDTH_MIN      = 10
WIDTH_MAX      = 60
SAMPLE_FOR_FIT = 1000

# Sprint 15 — ordem preferida de tabelas Protheus no relatorio
PREFERRED_TABLE_ORDER = (
    "SF1", "SD1", "SDS", "SDT",
    "F1",  "D1",  "DS",  "DT",      # variantes single-table (sem alias JOIN)
    "SA2", "SB1", "SA1", "SE2", "SDE", "SFT", "SF3", "CKO", "CK0",
)


def reorder_columns_by_prefix(cols: Sequence[str]) -> list[str]:
    """Reordena `cols` agrupando por prefixo Protheus (ate o `_` ou `__`)
    e aplicando a ordem preferida (SF1 -> SD1 -> SDS -> SDT)."""
    if not cols:
        return list(cols)

    def _prefix(c: str) -> str:
        if "__" in c:
            return c.split("__", 1)[0].upper()
        if "_" in c:
            return c.split("_", 1)[0].upper()
        return c.upper()

    groups: dict[str, list[str]] = {}
    fallback_order: list[str] = []
    for c in cols:
        p = _prefix(c)
        if p not in groups:
            groups[p] = []
            if p not in PREFERRED_TABLE_ORDER:
                fallback_order.append(p)
        groups[p].append(c)

    final_prefixes: list[str] = []
    seen: set[str] = set()
    for p in PREFERRED_TABLE_ORDER:
        if p in groups and p not in seen:
            final_prefixes.append(p); seen.add(p)
    for p in fallback_order:
        if p not in seen:
            final_prefixes.append(p); seen.add(p)
    return [c for p in final_prefixes for c in groups[p]]


def build_formatted_xlsx_bytes(
    rows: Iterable[dict],
    *,
    sheet_name: str = "Dados",
    columns: Optional[Sequence[str]] = None,
    reorder: bool = True,
) -> bytes:
    """Gera um XLSX formatado em memoria e devolve `bytes`.

    Args:
        rows:        iteravel de dicts (cada dict = uma linha).
        sheet_name:  nome da aba.
        columns:     ordem fixa das colunas. Se None, descobre na primeira row.
        reorder:     aplica `reorder_columns_by_prefix` (so se `columns is None`).

    Aceita rows vazias (gera planilha so com header se `columns` fornecido,
    ou uma celula placeholder se nao houver dados nem `columns`).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(
        start_color=HEADER_BG_HEX, end_color=HEADER_BG_HEX, fill_type="solid",
    )
    HEADER_FONT = Font(bold=True, color=HEADER_FG_HEX, size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows_list = list(rows or [])
    if columns is not None:
        ordered_cols = list(columns)
    else:
        if rows_list:
            raw = list(rows_list[0].keys())
            ordered_cols = reorder_columns_by_prefix(raw) if reorder else raw
        else:
            ordered_cols = ["(sem dados)"]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Dados"

    # 1) Header
    ws.append(ordered_cols)
    for col_idx in range(1, len(ordered_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # 2) AutoFit — sample com header + ate SAMPLE_FOR_FIT linhas
    widths: list[int] = []
    for i, col in enumerate(ordered_cols):
        max_len = len(str(col))
        for r in rows_list[:SAMPLE_FOR_FIT]:
            v = r.get(col) if isinstance(r, dict) else None
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        widths.append(min(max(max_len + 2, WIDTH_MIN), WIDTH_MAX))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 3) Dados
    for r in rows_list:
        ws.append([r.get(c) if isinstance(r, dict) else None for c in ordered_cols])

    # 4) Freeze + AutoFilter
    ws.freeze_panes = "A2"
    if rows_list:
        last_col = get_column_letter(len(ordered_cols))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows_list) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
