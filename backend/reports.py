"""Geracao de arquivos de relatorio.

Fase 3: alem das funcoes in-memory (uso historico), adicionamos
`write_xlsx_stream` e `write_csv_stream` que aceitam um iterador
de linhas (dict) e gravam direto no disco em modo streaming —
NUNCA carrega o dataset inteiro em memoria.

Uso pelo worker Celery:
    from .reports import write_xlsx_stream
    with engine.connect() as conn:
        gen = pd.read_sql(text(sql), conn, params=p, chunksize=10000)
        write_xlsx_stream(path, _iter_chunks_to_rows(gen), progress_cb)

Formatos suportados:
- xlsx (streaming via openpyxl write_only)
- csv  (streaming via csv.writer)
- ods  (in-memory, nao suporta streaming — limita a 100k linhas)
- pdf  (in-memory, nao adequado a > 50k linhas)
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Callable, Iterable, Iterator, Optional

import pandas as pd

from .config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()
OUTPUT_DIR = Path(settings.REPORTS_OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ---- Helpers comuns --------------------------------------------------------

# Hotfix Sprint 8 Part 4 — mensagem padrao gravada em arquivos quando a
# consulta retorna vazia ou inválida (ex: coluna inexistente). Antes o
# gerador estourava `NoneType has no len()`; agora produz um arquivo
# "amigavel" com 1 linha de aviso (operador entende sem suporte).
_EMPTY_PLACEHOLDER = [{
    "AVISO": "A consulta nao retornou dados ou as colunas solicitadas sao "
             "invalidas para esta tabela. Verifique o filtro e os nomes de coluna."
}]


def _to_dataframe(rows) -> pd.DataFrame:
    """Hotfix Sprint 8 Part 4: aceita None, list vazia ou iteravel normal.

    Antes:  `pd.DataFrame(list(rows) or [{}])` — `list(None)` lançava TypeError
            propagando como `NoneType has no len()` no openpyxl/pandas.
    Agora:  None / lista vazia → placeholder com mensagem amigavel.
    """
    if rows is None:
        return pd.DataFrame(_EMPTY_PLACEHOLDER)
    try:
        materialized = list(rows)
    except TypeError:
        return pd.DataFrame(_EMPTY_PLACEHOLDER)
    if not materialized:
        return pd.DataFrame(_EMPTY_PLACEHOLDER)
    return pd.DataFrame(materialized)


def _safe_name(name: str) -> str:
    return "".join(c for c in name if c.isalnum() or c in "._-") or "report"


# ============================================================================
# STREAMING — usado pelo worker Celery em relatorios grandes
# ============================================================================

def write_csv_stream(
    path: Path,
    row_iter: Iterator[dict],
    progress_cb: Optional[Callable[[int], None]] = None,
    cancel_cb: Optional[Callable[[], bool]] = None,
    include_physical: bool = False,
) -> int:
    """Grava CSV em disco linha a linha. Retorna numero de linhas escritas.

    `progress_cb(rows_so_far)` chamado a cada 1000 linhas (para o worker
    atualizar Job.progress_pct sem virar gargalo).
    `cancel_cb()` checada a cada 1000 linhas; True aborta e levanta `JobCanceled`.
    """
    rows_written = 0
    headers: Optional[list[str]] = None
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        for row in row_iter:
            if headers is None:
                headers = list(row.keys()) or ["(vazio)"]
                # Cabecalho humanizado (so titulo via SX3); dados continuam por
                # `headers` fisicos (lookup em row).
                try:
                    from . import dict_sx3
                    _fn = dict_sx3.humanize_header if include_physical else dict_sx3.humanize_title
                    writer.writerow([_fn(h) for h in headers])
                except Exception:
                    writer.writerow(headers)
            writer.writerow([_csv_value(row.get(h)) for h in headers])
            rows_written += 1
            if rows_written % 1000 == 0:
                if progress_cb: progress_cb(rows_written)
                if cancel_cb and cancel_cb():
                    raise JobCanceled("Cancelado pelo usuario")
    return rows_written


def write_xlsx_stream(
    path: Path,
    row_iter: Iterator[dict],
    progress_cb: Optional[Callable[[int], None]] = None,
    cancel_cb: Optional[Callable[[], bool]] = None,
    max_rows: int = 1_048_575,
    humanize_headers: bool = True,
    include_physical: bool = False,
) -> int:
    """Grava XLSX formatado em streaming (Sprint 5 + Sprint 6).

    Formatação aplicada:
    - Cabeçalho: fundo verde Fertimaxi (#2E8B3D) + texto branco bold.
    - **Sprint 6**: header traduzido via SX3 (`C5_EMISSAO` → `"Data de Emissão (C5_EMISSAO)"`).
      A tradução acontece APENAS no momento de escrever o header (1 lookup por
      coluna), nunca por linha — performance preservada.
    - Linha 1 congelada (freeze panes — usuário rola sem perder os títulos).
    - Larguras de coluna baseadas no cabeçalho + amostragem das primeiras
      `SAMPLE_FOR_WIDTH` linhas (compromisso entre AutoFit perfeito e streaming).
    - Limite Excel = 1.048.575 linhas. Acima disso → `RowLimitExceeded`.

    `humanize_headers=False` desativa a tradução SX3 (debug / testes).

    Trade-off: em `write_only` mode (memória O(1)), `column_dimensions` precisa
    ser setado ANTES do primeiro `append`. Para evitar carregar tudo em RAM,
    fazemos uma amostragem do começo: bufferizamos N linhas em memória,
    calculamos larguras, depois streamea o resto. Pior caso: a coluna fica
    estreita demais para um valor que apareceu DEPOIS da amostra — usuário
    arrasta manualmente. Para autofit 100%, use CSV ou dataset < 100k linhas
    em modo normal (não streaming).
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="FF2E8B3D", end_color="FF2E8B3D", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
    # Sprint 15: header centralizado (era 'left'), padronizando com o caminho in-memory
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    SAMPLE_FOR_WIDTH = 200  # primeiras N linhas usadas para autofit
    WIDTH_MIN, WIDTH_MAX = 10, 60

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Dados")
    ws.freeze_panes = "A2"

    rows_written = 0
    headers: Optional[list[str]] = None
    sample_buffer: list[dict] = []
    col_widths: dict[int, int] = {}

    def _track(col_idx: int, value) -> None:
        s = "" if value is None else str(value)
        n = min(len(s), WIDTH_MAX)
        if n > col_widths.get(col_idx, 0):
            col_widths[col_idx] = n

    def _flush_header_and_sample():
        """Escreve cabeçalho formatado + amostra bufferizada, após calcular widths.

        Sprint 6: a tradução SX3 acontece AQUI (1 lookup por coluna, zero overhead
        por linha de dados). headers já foi descoberto pela primeira row.
        """
        # 0) Traduz headers via SX3 — momento certo: headers ja conhecidos
        if humanize_headers:
            try:
                from . import dict_sx3
                _fn = dict_sx3.humanize_header if include_physical else dict_sx3.humanize_title
                display_headers = [_fn(h) for h in headers]
            except Exception:
                display_headers = list(headers)
        else:
            display_headers = list(headers)

        # 1) Larguras das colunas — setar ANTES de qualquer append (write_only)
        for i, _h in enumerate(headers):
            human = display_headers[i]
            content_len = max(col_widths.get(i, WIDTH_MIN), len(human))
            width = max(WIDTH_MIN, min(content_len + 2, WIDTH_MAX))
            ws.column_dimensions[get_column_letter(i + 1)].width = width
        # 2) Cabeçalho formatado (traduzido via SX3 quando humanize_headers=True)
        header_cells = []
        for h_display in display_headers:
            cell = WriteOnlyCell(ws, value=h_display)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            header_cells.append(cell)
        ws.append(header_cells)
        # 3) Amostra — usa headers ORIGINAIS para fazer lookup nos dicts (`row[h]`)
        for r in sample_buffer:
            ws.append([r.get(h) for h in headers])

    try:
        for row in row_iter:
            # FASE 1: bufferiza ate SAMPLE_FOR_WIDTH para calcular larguras
            if headers is None:
                # Sprint 15: aplica ordem preferida SF1 -> SD1 -> SDS -> SDT
                raw_headers = list(row.keys()) or ["(vazio)"]
                headers = _reorder_columns_by_prefix(raw_headers)
                for i, h in enumerate(headers):
                    _track(i, h)

            if len(sample_buffer) < SAMPLE_FOR_WIDTH:
                for i, h in enumerate(headers):
                    _track(i, row.get(h))
                sample_buffer.append(row)
                rows_written += 1
                if cancel_cb and cancel_cb():
                    raise JobCanceled("Cancelado pelo usuario")
                continue

            # FASE 2: amostra completa — flush e streamea o resto
            if sample_buffer is not None:
                _flush_header_and_sample()
                sample_buffer = None  # marca como flushado

            if rows_written >= max_rows:
                raise RowLimitExceeded(
                    f"Limite XLSX atingido ({max_rows}). Use CSV para datasets maiores."
                )
            ws.append([row.get(h) for h in headers])
            rows_written += 1
            if rows_written % 1000 == 0:
                if progress_cb: progress_cb(rows_written)
                if cancel_cb and cancel_cb():
                    raise JobCanceled("Cancelado pelo usuario")

        # Caso o dataset SEJA menor que a amostra → flush ao final
        if sample_buffer is not None and headers is not None:
            _flush_header_and_sample()
        # Se nem headers conseguimos (iterator vazio), cria placeholder
        if headers is None:
            ws.append(["(sem dados)"])

        wb.save(path)
    finally:
        wb.close()
    return rows_written


def _csv_value(v):
    if v is None: return ""
    if isinstance(v, float):
        # Evita "1.0" virar "1" em CSV BR
        return f"{v}".replace(".", ",") if v % 1 else f"{int(v)}"
    return str(v)


class JobCanceled(RuntimeError):
    """Sinaliza cancelamento cooperativo."""


class RowLimitExceeded(RuntimeError):
    """Dataset excede o limite do formato escolhido."""


# ============================================================================
# IN-MEMORY (legacy — usado quando rows estimado < threshold)
# ============================================================================

def to_xlsx(rows: Iterable[dict], path: Path) -> Path:
    df = _to_dataframe(rows)
    df.to_excel(path, index=False, engine="openpyxl")
    return path


def to_csv(rows: Iterable[dict], path: Path) -> Path:
    df = _to_dataframe(rows)
    df.to_csv(path, index=False, sep=";", quoting=csv.QUOTE_MINIMAL, encoding="utf-8-sig")
    return path


def to_ods(rows: Iterable[dict], path: Path) -> Path:
    df = _to_dataframe(rows)
    df.to_excel(path, index=False, engine="odf")
    return path


def to_pdf(rows: Iterable[dict], path: Path, title: str = "Relatorio Protheus") -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    rows = list(rows)
    headers = list(rows[0].keys()) if rows else ["(sem dados)"]
    data = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E8B3D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)
    doc.build(elements)
    return path


def generate(rows: Iterable[dict], file_format: str, base_name: str) -> Path:
    file_format = file_format.lower()
    target = OUTPUT_DIR / f"{_safe_name(base_name)}.{file_format}"
    if file_format == "xlsx": return to_xlsx(rows, target)
    if file_format == "csv":  return to_csv(rows, target)
    if file_format == "ods":  return to_ods(rows, target)
    if file_format == "pdf":  return to_pdf(rows, target, title=base_name)
    raise ValueError(f"Formato nao suportado: {file_format}")


# Sprint 15 — ordem preferida de tabelas no relatorio Excel.
# Diretiva da diretoria: SF1 (cab. fiscal) -> SD1 (itens fiscal) -> SDS (XML cab)
# -> SDT (XML itens) -> auxiliares (SA2, SB1, ...) -> outras na ordem de aparicao.
PREFERRED_TABLE_ORDER = (
    "SF1", "SD1", "SDS", "SDT",
    "F1",  "D1",  "DS",  "DT",     # variantes single-table (sem alias)
    "SA2", "SB1", "SA1", "SE2", "SDE", "SFT", "SF3", "CKO", "CK0",
)


def _reorder_columns_by_prefix(cols: list[str]) -> list[str]:
    """Reordena colunas agrupando pelo prefixo da tabela.

    Aceita 2 formatos:
      - JOIN qualified (Sprint 5): `SF1__F1_DOC`, `SD1__D1_COD`
      - Single-table:              `F1_DOC`, `D1_COD`, `DS_DOC`, `DT_QUANT`

    Sprint 15: aplica a ordem PREFERIDA `PREFERRED_TABLE_ORDER` primeiro
    (SF1 -> SD1 -> SDS -> SDT). Prefixos nao listados vao ao final na ordem
    de aparicao. Estavel dentro de cada grupo.
    """
    if not cols:
        return cols

    def _prefix(c: str) -> str:
        if "__" in c:
            return c.split("__", 1)[0].upper()
        if "_" in c:
            return c.split("_", 1)[0].upper()
        return c.upper()

    groups: dict[str, list[str]] = {}
    fallback_order: list[str] = []
    for c in cols:
        p = _prefix(c)
        if p not in groups:
            groups[p] = []
            if p not in PREFERRED_TABLE_ORDER:
                fallback_order.append(p)
        groups[p].append(c)

    final_prefixes: list[str] = []
    seen: set[str] = set()
    # 1) Prefixos preferidos (na ordem fixa) — so se aparecerem
    for p in PREFERRED_TABLE_ORDER:
        if p in groups and p not in seen:
            final_prefixes.append(p)
            seen.add(p)
    # 2) Demais prefixos na ordem de aparicao
    for p in fallback_order:
        if p not in seen:
            final_prefixes.append(p)
            seen.add(p)
    return [c for p in final_prefixes for c in groups[p]]


def _write_formatted_xlsx(df: pd.DataFrame, buf: io.BytesIO) -> None:
    """Sprint 11 — XLSX com header verde Fertimaxi + AutoFit + freeze panes.

    Replica a formatacao do `write_xlsx_stream` (que ja existia para o
    worker Celery) para o caminho sincrono usado por `/api/protheus/download`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="FF2E8B3D", end_color="FF2E8B3D", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WIDTH_MIN, WIDTH_MAX = 10, 60

    # Reordena colunas agrupadas por prefixo da tabela
    ordered_cols = _reorder_columns_by_prefix(list(df.columns))
    df = df[ordered_cols] if ordered_cols else df

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # 1) Header com formatacao Fertimaxi
    ws.append(ordered_cols)
    for col_idx in range(1, len(ordered_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # 2) AutoFit — sample com toda a coluna (modo in-memory)
    # Para cada coluna, calcula a maior largura (cap WIDTH_MAX) baseada nos
    # valores reais + o header.
    widths: list[int] = []
    for i, col_name in enumerate(ordered_cols):
        max_len = len(str(col_name))
        series = df[col_name]
        # Limita o sample a 1000 linhas para nao engasgar com datasets enormes
        for v in series.head(1000):
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        widths.append(min(max(max_len + 2, WIDTH_MIN), WIDTH_MAX))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 3) Dados — bulk append (rapido para datasets ate ~50k linhas)
    for _, row in df.iterrows():
        ws.append([row.get(c) for c in ordered_cols])

    # 4) Freeze panes (linha 1 fixa ao rolar) + auto-filter
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(ordered_cols))}{len(df) + 1}"

    wb.save(buf)
    wb.close()


def _human_columns(cols: list, include_physical: bool = False) -> dict:
    """Mapa {coluna_fisica: rotulo} para renomear cabecalhos no export.
    - include_physical=False (operador): SO o titulo SX3.
    - include_physical=True  (admin):    "Titulo (FISICO)".
    Mantem o fisico se a SX3 nao tiver titulo; desambigua com o fisico entre
    parenteses se dois campos mapearem para o mesmo titulo (evita coluna
    duplicada no pandas/openpyxl)."""
    from . import dict_sx3
    used: set = set()
    out: dict = {}
    for c in cols:
        t = dict_sx3.humanize_title(c)
        if not t or t == c:
            out[c] = c
            continue
        label = f"{t} ({c})" if include_physical else t
        if label in used:
            label = f"{t} ({c})"
        used.add(label)
        out[c] = label
    return out


def to_bytes(rows: Iterable[dict], file_format: str,
             include_physical: bool = False) -> tuple[bytes, str, str]:
    buf = io.BytesIO()
    df = _to_dataframe(rows)
    # Reordena por prefixo FISICO e humaniza os cabecalhos UMA vez, para todos os
    # formatos (xlsx/csv/ods/pdf). As reordenacoes seguintes operam sobre os
    # rotulos humanizados (sem prefixo) e viram no-op.
    _ordered = _reorder_columns_by_prefix(list(df.columns))
    if _ordered:
        df = df[_ordered]
    df = df.rename(columns=_human_columns(list(df.columns), include_physical))
    file_format = file_format.lower()
    if file_format == "xlsx":
        # Sprint 11 — XLSX formatado (verde Fertimaxi + AutoFit + ordenacao)
        _write_formatted_xlsx(df, buf)
        return (
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if file_format == "csv":
        # Sprint 11 — CSV tambem reordena colunas pelo prefixo
        ordered_cols = _reorder_columns_by_prefix(list(df.columns))
        df = df[ordered_cols] if ordered_cols else df
        text = df.to_csv(index=False, sep=";", quoting=csv.QUOTE_MINIMAL, encoding="utf-8-sig")
        return text.encode("utf-8-sig"), "text/csv; charset=utf-8", "csv"
    if file_format == "ods":
        ordered_cols = _reorder_columns_by_prefix(list(df.columns))
        df = df[ordered_cols] if ordered_cols else df
        with pd.ExcelWriter(buf, engine="odf") as writer:
            df.to_excel(writer, index=False)
        return buf.getvalue(), "application/vnd.oasis.opendocument.spreadsheet", "ods"
    if file_format == "pdf":
        tmp = OUTPUT_DIR / "_preview.pdf"
        to_pdf(df.to_dict(orient="records"), tmp)
        data = tmp.read_bytes()
        tmp.unlink(missing_ok=True)
        return data, "application/pdf", "pdf"
    raise ValueError(f"Formato nao suportado: {file_format}")
